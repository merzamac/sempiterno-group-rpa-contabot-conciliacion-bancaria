# movimientos/excel_builder.py
from abc import ABC, abstractmethod
import re
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from .movimiento_strategy import MovimientoStrategy
from contabot_conciliacion_bancaria.process.conciliacion.app.factory import (
    get_header_model,
)
from contabot_conciliacion_bancaria.process.conciliacion.app.header import HeaderMasivo


class ExcelBuilder(ABC):
    def __init__(self):
        self.wb = Workbook()
        self._remove_default_sheet()

    def _remove_default_sheet(self):
        if self.wb.sheetnames:
            self.wb.remove(self.wb.active)

    def build(self) -> Workbook:
        return self.wb


class ReporteExcelBuilder2(ExcelBuilder):
    def __init__(self):
        super().__init__()
        self.fill_amarillo = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        self.fill_azul = PatternFill(
            start_color="1c55a7", end_color="1c55a7", fill_type="solid"
        )

    def agregar_sheet(self, nombre: str, datos: dict):
        ws = self.wb.create_sheet(title=nombre)
        header = nombre.split(" ")[1]
        model_header = get_header_model(header)
        # Configurar header
        self._configurar_header(ws, model_header.values())

        # Agregar datos
        self._agregar_datos(ws, datos["movimientos"])

        # Resaltar coincidencias
        self._resaltar_coincidencias(ws, datos["coincidencias"])

        # Ajustar columnas
        self._ajustar_columnas(ws)

    def _configurar_header(self, ws, headers: tuple):
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=1, column=col, value=header)
            celda.fill = self.fill_azul

    def _agregar_datos(self, ws, movimientos: list):

        for i, movimiento in enumerate(movimientos, 2):
            # Convertir movimiento a tupla para append
            ws.append(tuple(movimiento))

    # def _movimiento_a_tupla(self, movimiento):
    #     # Convertir objeto movimiento a tupla según su estructura
    #     if hasattr(movimiento, "__dict__"):
    #         return tuple(vars(movimiento).values())
    #     return tuple(movimiento)

    def _resaltar_coincidencias(self, ws, filas_coincidentes: list[int]):
        if not filas_coincidentes:
            return

        num_columnas = ws.max_column
        for num_fila in filas_coincidentes:
            for col in range(1, num_columnas + 1):
                celda = ws.cell(row=num_fila, column=col)
                celda.fill = self.fill_amarillo

    def _ajustar_columnas(self, ws):
        for col in range(1, ws.max_column + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                try:
                    value = ws.cell(row=row, column=col).value
                    if value:
                        max_length = max(max_length, len(str(value)))
                except:
                    pass
            col_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[col_letter].width = max_length + 2


class MasivoExcelBuilder2(ExcelBuilder):
    def __init__(self):
        super().__init__()

    def _configurar_header(self, ws, headers: tuple):
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=1, column=col, value=header)
            # celda.fill = self.fill_azul

    def agregar_sheet(self, nombre: str, datos: list):
        ws = self.wb.create_sheet(title=nombre)
        # Configurar header
        self._configurar_header(ws, HeaderMasivo.values())

        # Agregar datos
        self._agregar_datos(ws, datos)

        # Ajustar columnas
        self._ajustar_columnas(ws)

    def _agregar_datos(self, ws, datos: list):

        for movimiento in datos:
            # Convertir movimiento a tupla para append
            ws.append(tuple(movimiento))

    def _ajustar_columnas(self, ws):
        for col in range(1, ws.max_column + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                try:
                    value = ws.cell(row=row, column=col).value
                    if value:
                        max_length = max(max_length, len(str(value)))
                except:
                    pass
            col_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[col_letter].width = max_length + 2


class ExcelOpenpyxl:
    def __init__(self):
        self.wb = Workbook()
        self._remove_default_sheet()

    def _remove_default_sheet(self):
        if self.wb.sheetnames:
            self.wb.remove(self.wb.active)

    def build(self) -> Workbook:
        return self.wb

    def create_sheet(self, name_sheet: str) -> Worksheet:
        return self.wb.create_sheet(title=name_sheet)

    def add_header(
        self, ws: Worksheet, header: tuple, hex_color: str = "FFFFFF"
    ) -> None:
        color = PatternFill(
            start_color=hex_color, end_color=hex_color, fill_type="solid"
        )
        for column, value_column in enumerate(header, 1):
            cell_fill = ws.cell(row=1, column=column, value=value_column)
            cell_fill.fill = color

    def add_data(self, ws, data: tuple):

        for row in data:
            # Convertir movimiento a tupla para append
            ws.append(tuple(row))

    def highlight_rows(self, ws, rows: tuple, hex_color: str = "FFFF00"):
        if not rows:
            return

        color = PatternFill(
            start_color=hex_color, end_color=hex_color, fill_type="solid"
        )

        num_columnas = ws.max_column
        for n_rows in rows:
            for col in range(1, num_columnas + 1):
                cell_fill = ws.cell(row=n_rows, column=col)
                cell_fill.fill = color

    def fix_column_size(self, ws):
        for col in range(1, ws.max_column + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                try:
                    value = ws.cell(row=row, column=col).value
                    if value:
                        max_length = max(max_length, len(str(value)))
                except:
                    pass
            col_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[col_letter].width = max_length + 2


class ReportExcelBuilder(ExcelOpenpyxl):
    def __init__(self):
        super().__init__()

    def make_report(self, nombre: str, datos: dict):
        ws = self.create_sheet(nombre)
        header = nombre.split(" ")[1]
        model_header = get_header_model(header)
        # Configurar header
        self.add_header(ws, model_header.values(), "1c55a7")

        # Agregar datos
        self.add_data(ws, tuple(datos["movimientos"]))

        # Resaltar coincidencias
        self.highlight_rows(ws, tuple(datos["coincidencias"]))

        # Ajustar columnas
        self.fix_column_size(ws)


class MasivoExcelBuilder(ExcelOpenpyxl):
    def __init__(self):
        super().__init__()

    def make_report(self, name_sheet: str, data: list):
        ws = self.create_sheet(name_sheet)
        # Configurar header
        self.add_header(ws, HeaderMasivo.values(), "cc1758")

        # Agregar datos
        self.add_data(ws, tuple(data))

        # Ajustar columnas
        self.fix_column_size(ws)
